"""Regression tests for recurring-bill export normalization."""

from __future__ import annotations

from pathlib import Path
import runpy

import duckdb
from openpyxl import load_workbook
import pandas as pd

from data.recurring_bills.export_resolution import classify_recurrence_type


_REPO_ROOT = Path(__file__).resolve().parent.parent
_EXPORT_ALL = runpy.run_path(
    str(_REPO_ROOT / "scripts" / "export_all_bills_classified.py")
)["export"]
_EXPORT_OUR_SCAN = runpy.run_path(
    str(_REPO_ROOT / "scripts" / "export_amnon_from_our_scan.py")
)["export"]
_REQUIRED_EXPORT_SHEETS = [
    "Classified Bills",
    "Reference Resolution",
    "Data Dictionary",
]


def _build_db(db_path: Path) -> None:
    con = duckdb.connect(str(db_path))
    try:
        con.execute(
            """
            CREATE TABLE KNS_Bill (
                BillID BIGINT,
                KnessetNum BIGINT,
                Name VARCHAR,
                PrivateNumber BIGINT,
                SubTypeDesc VARCHAR
            )
            """
        )
        con.executemany(
            """
            INSERT INTO KNS_Bill (BillID, KnessetNum, Name, PrivateNumber, SubTypeDesc)
            VALUES (?, ?, ?, ?, ?)
            """,
            [
                (1, 18, "Original bill", 111, "Private"),
                (2, 19, "Current bill", 222, "Private"),
                (3, 19, "Mid-chain bill", 333, "Private"),
                (4, 1, "Historic bill", 444, "Private"),
                (5, 19, "Plural recurring bill", 555, "Private"),
                (6, 16, "Linked to unresolved parent", 666, "Private"),
                (7, 15, "Unresolved parent", 777, "Private"),
                (8, 16, "Ambiguous multi-reference bill", 888, "Private"),
                (9, 15, "Similar bill one", 1396, "Private"),
                (10, 15, "Similar bill two", 3512, "Private"),
                (
                    11,
                    11,
                    "Source old-Knesset phrase bill",
                    274,
                    "Private",
                ),
                (
                    12,
                    11,
                    "Same-name source-Knesset bill",
                    294,
                    "Private",
                ),
            ],
        )
        con.execute(
            """
            CREATE TABLE bill_classifications_doc_full (
                BillID BIGINT,
                KnessetNum BIGINT,
                Name VARCHAR,
                PrivateNumber BIGINT,
                is_original BOOLEAN,
                original_bill_id BIGINT,
                matched_phrase VARCHAR,
                method VARCHAR,
                reference_candidates VARCHAR,
                reference_candidate_count BIGINT,
                reference_resolution_reason VARCHAR,
                reference_resolution_confidence DOUBLE,
                multiple_references_detected BOOLEAN,
                submission_date VARCHAR,
                suspicious_self_resolution BOOLEAN,
                ambiguous_reference_resolution BOOLEAN,
                ambiguous_reference_reason VARCHAR,
                doc_url VARCHAR,
                classification_source VARCHAR,
                last_updated TIMESTAMP
            )
            """
        )
        con.executemany(
            """
            INSERT INTO bill_classifications_doc_full VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    1,
                    18,
                    "Original bill",
                    111,
                    True,
                    1,
                    None,
                    "doc_no_pattern",
                    "[]",
                    0,
                    None,
                    None,
                    False,
                    "2013-01-01",
                    False,
                    False,
                    None,
                    "https://example/1.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    2,
                    19,
                    "Current bill",
                    222,
                    False,
                    3,
                    "הצעת חוק זהה",
                    "doc_pattern_linked",
                    "[]",
                    1,
                    "explicit_private_number_and_knesset",
                    0.99,
                    False,
                    "2014-01-15",
                    False,
                    False,
                    None,
                    "https://example/2.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    3,
                    19,
                    "Mid-chain bill",
                    333,
                    False,
                    1,
                    "הצעת חוק דומה",
                    "doc_pattern_linked",
                    """[{"resolved_bill_id": 1, "private_number": 111, "referenced_knesset": 18, "reference_text": "פ/111/18", "reference_resolution_reason": "explicit_private_number_and_knesset", "reference_resolution_confidence": 0.99, "selected": true, "suspicious_self_resolution": false}]""",
                    1,
                    "explicit_private_number_and_knesset",
                    0.99,
                    False,
                    "2014-01-10",
                    False,
                    False,
                    None,
                    "https://example/3.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    4,
                    1,
                    "Historic bill",
                    444,
                    True,
                    4,
                    None,
                    "doc_no_pattern",
                    "[]",
                    0,
                    None,
                    None,
                    False,
                    "2049-11-23",
                    False,
                    False,
                    None,
                    "https://example/4.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    5,
                    19,
                    "Plural recurring bill",
                    555,
                    False,
                    1,
                    "הצעות  חוק  דומות  בעיקרן",
                    "doc_pattern_linked",
                    "[]",
                    1,
                    "explicit_private_number_and_knesset",
                    0.99,
                    False,
                    "2014-02-24",
                    False,
                    False,
                    None,
                    "https://example/5.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    6,
                    16,
                    "Linked to unresolved parent",
                    666,
                    False,
                    7,
                    "הצעת חוק זהה",
                    "doc_pattern_linked",
                    """[{"resolved_bill_id": 7, "private_number": 777, "referenced_knesset": 15, "reference_text": "פ/777", "reference_resolution_reason": "contextual_knesset_phrase_match", "reference_resolution_confidence": 0.92, "selected": true, "suspicious_self_resolution": false}]""",
                    1,
                    "contextual_knesset_phrase_match",
                    0.92,
                    False,
                    "2004-05-17",
                    False,
                    False,
                    None,
                    "https://example/6.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    7,
                    15,
                    "Unresolved parent",
                    777,
                    False,
                    7,
                    "הצעת חוק דומה",
                    "doc_pattern_unresolved",
                    "[]",
                    0,
                    "no_reference_candidates_in_recurrence_context",
                    None,
                    False,
                    "2000-01-01",
                    False,
                    False,
                    None,
                    "https://example/7.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    8,
                    16,
                    "Ambiguous multi-reference bill",
                    888,
                    False,
                    8,
                    "הצעת חוק דומה",
                    "doc_pattern_unresolved",
                    """[{"resolved_bill_id": 9, "private_number": 1396, "referenced_knesset": 15, "reference_text": "פ/1396", "reference_resolution_reason": "contextual_knesset_phrase_match", "reference_resolution_confidence": 0.92, "selected": false, "suspicious_self_resolution": false, "tied_for_best": true}, {"resolved_bill_id": 10, "private_number": 3512, "referenced_knesset": 15, "reference_text": "פ/3512", "reference_resolution_reason": "contextual_knesset_phrase_match", "reference_resolution_confidence": 0.92, "selected": false, "suspicious_self_resolution": false, "tied_for_best": true}]""",
                    2,
                    "ambiguous_primary_reference_candidates",
                    None,
                    True,
                    "2004-05-17",
                    False,
                    True,
                    "multiple_equally_strong_candidates",
                    "https://example/8.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    9,
                    15,
                    "Similar bill one",
                    1396,
                    True,
                    9,
                    None,
                    "doc_no_pattern",
                    "[]",
                    0,
                    None,
                    None,
                    False,
                    "2000-01-01",
                    False,
                    False,
                    None,
                    "https://example/9.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    10,
                    15,
                    "Similar bill two",
                    3512,
                    True,
                    10,
                    None,
                    "doc_no_pattern",
                    "[]",
                    0,
                    None,
                    None,
                    False,
                    "2000-01-01",
                    False,
                    False,
                    None,
                    "https://example/10.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    11,
                    11,
                    "Source old-Knesset phrase bill",
                    274,
                    False,
                    12,
                    "הצעת חוק זהה",
                    "doc_pattern_linked",
                    """[{"context": "הצעת חוק זהה הוגשה בכנסת העשירית ע\\"י חברי הכנסת", "contextual_knesset": 10, "explicit_knesset": null, "phrase_index": 0, "phrase_text": "הצעת חוק זהה", "priority": 2, "private_number": null, "recurrence_type": "identical", "reference_index": -1, "reference_resolution_confidence": 0.58, "reference_resolution_reason": "same_knesset_name_fallback", "reference_text": null, "referenced_knesset": 10, "resolved_bill_id": 12, "selected": true, "selection_rank": [2, 1, 0, -1000000], "suspicious_self_resolution": false, "tied_for_best": false}]""",
                    1,
                    "same_knesset_name_fallback",
                    0.58,
                    False,
                    "1986-02-17",
                    False,
                    False,
                    None,
                    "https://fs.knesset.gov.il//11/law/11_lst_534232.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
                (
                    12,
                    11,
                    "Same-name source-Knesset bill",
                    294,
                    True,
                    12,
                    None,
                    "doc_no_pattern",
                    "[]",
                    0,
                    None,
                    None,
                    False,
                    "1986-02-17",
                    False,
                    False,
                    None,
                    "https://example/12.doc",
                    "doc_based_full",
                    pd.Timestamp("2026-04-23 00:00:00"),
                ),
            ],
        )
    finally:
        con.close()


class TestExportResolution:
    def test_export_scripts_share_final_original_enrichment(self, tmp_path: Path):
        db_path = tmp_path / "warehouse.duckdb"
        excel_path = tmp_path / "amnon.xlsx"
        output_scan = tmp_path / "our_scan.xlsx"
        output_all = tmp_path / "all.xlsx"

        _build_db(db_path)
        pd.DataFrame(
            [
                {"BillID": 1, "Name": "Original bill"},
                {"BillID": 2, "Name": "Current bill"},
                {"BillID": 3, "Name": "Mid-chain bill"},
                {"BillID": 4, "Name": "Historic bill"},
                {"BillID": 5, "Name": "Plural recurring bill"},
            ]
        ).to_excel(excel_path, index=False)

        _EXPORT_OUR_SCAN(
            excel_path=excel_path, db_path=db_path, output_path=output_scan
        )
        _EXPORT_ALL(db_path=db_path, output_path=output_all)

        df_scan = pd.read_excel(output_scan)
        df_all = pd.read_excel(output_all)

        row_scan = df_scan.loc[df_scan["BillID"] == 2].iloc[0]
        row_all = df_all.loc[df_all["BillID"] == 2].iloc[0]

        assert row_scan["original_bill_id"] == 1
        assert row_all["effective_original_bill_id"] == 1
        assert row_scan["original_knesset_num"] == 18
        assert row_all["effective_original_knesset_num"] == 18
        assert row_scan["original_private_number"] == 111
        assert row_all["effective_original_private_number"] == 111
        assert row_scan["submission_date"] == "2014-01-15"
        assert row_all["submission_date"] == "2014-01-15"

    def test_export_scripts_drop_implausible_submission_dates(self, tmp_path: Path):
        db_path = tmp_path / "warehouse.duckdb"
        excel_path = tmp_path / "amnon.xlsx"
        output_scan = tmp_path / "our_scan.xlsx"
        output_all = tmp_path / "all.xlsx"

        _build_db(db_path)
        pd.DataFrame(
            [
                {"BillID": 1, "Name": "Original bill"},
                {"BillID": 2, "Name": "Current bill"},
                {"BillID": 3, "Name": "Mid-chain bill"},
                {"BillID": 4, "Name": "Historic bill"},
                {"BillID": 5, "Name": "Plural recurring bill"},
            ]
        ).to_excel(excel_path, index=False)

        _EXPORT_OUR_SCAN(
            excel_path=excel_path, db_path=db_path, output_path=output_scan
        )
        _EXPORT_ALL(db_path=db_path, output_path=output_all)

        df_scan = pd.read_excel(output_scan)
        df_all = pd.read_excel(output_all)

        row_scan = df_scan.loc[df_scan["BillID"] == 4].iloc[0]
        row_all = df_all.loc[df_all["BillID"] == 4].iloc[0]

        assert pd.isna(row_scan["submission_date"])
        assert pd.isna(row_all["submission_date"])

    def test_classify_recurrence_type_handles_plural_variants(self):
        assert classify_recurrence_type("הצעות חוק זהות") == "identical"
        assert classify_recurrence_type("הצעות חוק דומות") == "similar"
        assert classify_recurrence_type("הצעות חוק דומות בעיקרן") == "similar"
        assert classify_recurrence_type("הצעות  חוק  דומות  בעיקרן") == "similar"

    def test_final_export_writes_corrected_submission_date_and_explicit_relation_type(
        self, tmp_path: Path
    ):
        db_path = tmp_path / "warehouse.duckdb"
        output_all = tmp_path / "all.xlsx"

        _build_db(db_path)
        _EXPORT_ALL(db_path=db_path, output_path=output_all)

        df_all = pd.read_excel(output_all)
        historic_row = df_all.loc[df_all["BillID"] == 4].iloc[0]
        plural_row = df_all.loc[df_all["BillID"] == 5].iloc[0]

        assert pd.isna(historic_row["submission_date"])
        assert plural_row["submission_date"] == "2014-02-24"
        assert plural_row["explicit_relation_type"] == "similar"

    def test_final_export_preserves_direct_reference_when_effective_origin_is_unresolved(
        self, tmp_path: Path
    ):
        db_path = tmp_path / "warehouse.duckdb"
        output_all = tmp_path / "all.xlsx"

        _build_db(db_path)
        _EXPORT_ALL(db_path=db_path, output_path=output_all)

        df_all = pd.read_excel(output_all)
        row = df_all.loc[df_all["BillID"] == 6].iloc[0]

        assert bool(row["is_recurring_upstream"]) is True
        assert row["direct_reference_bill_id"] == 7
        assert row["direct_reference_knesset_num"] == 15
        assert row["direct_reference_private_number"] == 777
        assert row["direct_reference"] == "15/777"
        assert row["cited_references"] == "15/777"

    def test_final_export_surfaces_all_resolved_reference_candidates(
        self, tmp_path: Path
    ):
        db_path = tmp_path / "warehouse.duckdb"
        output_all = tmp_path / "all.xlsx"

        _build_db(db_path)
        _EXPORT_ALL(db_path=db_path, output_path=output_all)

        df_all = pd.read_excel(output_all)
        row = df_all.loc[df_all["BillID"] == 8].iloc[0]

        assert bool(row["ambiguous_reference_resolution"]) is True
        assert pd.isna(row["direct_reference_bill_id"])
        assert row["cited_reference_count"] == 2
        assert row["cited_bill_ids"] == "9; 10"
        assert row["cited_references"] == "15/1396; 15/3512"

    def test_final_export_writes_reference_resolution_sheet_with_one_row_per_reference(
        self, tmp_path: Path
    ):
        db_path = tmp_path / "warehouse.duckdb"
        output_all = tmp_path / "all.xlsx"

        _build_db(db_path)
        _EXPORT_ALL(db_path=db_path, output_path=output_all)

        wb = load_workbook(output_all, read_only=True, data_only=True)
        try:
            assert wb.sheetnames == _REQUIRED_EXPORT_SHEETS
        finally:
            wb.close()

        workbook = pd.ExcelFile(output_all)
        assert workbook.sheet_names == _REQUIRED_EXPORT_SHEETS

        classified = pd.read_excel(workbook, sheet_name="Classified Bills")
        refs = pd.read_excel(workbook, sheet_name="Reference Resolution")
        dictionary = pd.read_excel(workbook, sheet_name="Data Dictionary")
        dictionary_columns = set(dictionary["column"].dropna())
        assert set(classified.columns).issubset(dictionary_columns)
        assert set(refs.columns).issubset(dictionary_columns)
        assert "same_knesset_name_fallback" not in set(
            refs["target_resolution_method"].dropna()
        )
        assert "same_knesset_name_fallback" not in set(
            classified["target_resolution_method"].dropna()
        )

        ambiguous_refs = refs.loc[refs["source_bill_id"] == 8]
        assert ambiguous_refs["reference_index"].tolist() == [1, 2]
        assert ambiguous_refs["target_bill_number_extracted"].tolist() == [1396, 3512]

        professor_example = refs.loc[refs["source_bill_id"] == 11].iloc[0]
        assert professor_example["source_knesset"] == 11
        assert professor_example["source_doc_id"] == 534232
        assert professor_example["target_knesset_extracted"] == 10
        assert pd.isna(professor_example["target_bill_number_extracted"])
        assert pd.isna(professor_example["target_url_extracted"])
        assert (
            professor_example["target_resolution_status"]
            == "unresolved_no_link_or_number"
        )
        assert pd.isna(professor_example["target_resolution_confidence"])

        source_row = classified.loc[classified["source_bill_id"] == 11].iloc[0]
        assert source_row["target_resolution_status"] == "unresolved_no_link_or_number"
        assert pd.isna(source_row["direct_reference_bill_id"])
        assert pd.isna(source_row["direct_reference_knesset_num"])
        assert pd.isna(source_row["direct_reference_private_number"])
        assert pd.isna(source_row["direct_reference"])
        assert bool(source_row["is_effective_original"]) is True
        assert source_row["effective_original_bill_id"] == 11
        assert source_row["effective_original_knesset_num"] == 11
        assert source_row["effective_original_private_number"] == 274
